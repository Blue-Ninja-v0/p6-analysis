import pandas as pd
import os
import sys
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

def clean_date(date_str):
    """Convert a date string to a datetime.date object."""
    if pd.isna(date_str):
        return None
    try:
        return pd.to_datetime(date_str).date()
    except ValueError:
        return None

def validate_and_clean_data(df):
    """Clean and validate the input DataFrame."""
    date_columns = ['target_start_date', 'target_end_date', 'act_start_date', 'act_end_date', 'early_start_date', 'early_end_date', 'late_start_date', 'late_end_date']
    for col in date_columns:
        if col in df.columns:
            df[col] = df[col].apply(clean_date)
    
    df['planned_duration'] = pd.to_numeric(df['planned_duration'], errors='coerce').fillna(0)
    
    return df

def get_column_names(df):
    """Map the DataFrame columns to expected column names."""
    column_mapping = {
        'task_id': ['task_id', 'activity_id', 'Task ID'],
        'task_name': ['task_name', 'activity_name', 'Task Name'],
        'wbs_id': ['wbs_id', 'WBS ID'],
        'wbs_name': ['wbs_name', 'WBS Name'],
        'planned_duration': ['planned_duration', 'target_drtn_hr_cnt', 'Duration'],
        'start_date': ['target_start_date', 'early_start_date', 'act_start_date'],
        'end_date': ['target_end_date', 'early_end_date', 'act_end_date'],
        'predecessors': ['pred_task_id', 'predecessors', 'predecessor_task_ids'],
        'phys_complete_pct': ['phys_complete_pct', '% complete'],
        'task_type': ['task_type', 'type'],
        'task_code': ['task_code', 'task code'],
        'rsrc_name': ['rsrc_name', 'resource'],
        'actv_code_name': ['actv_code_name', 'alt WBS']
    }
    
    selected_columns = {}
    for key, options in column_mapping.items():
        for option in options:
            if option in df.columns:
                selected_columns[key] = option
                break
        if key not in selected_columns and key not in ['predecessors', 'phys_complete_pct', 'task_type', 'task_code', 'rsrc_name', 'actv_code_name']:
            logging.warning(f"Could not find a column for {key}")
    
    return selected_columns

def load_data(file_path):
    """Load data from the specified Excel file."""
    try:
        main_df = pd.read_excel(file_path, sheet_name='Enhanced Baseline Data')
        cp_df = pd.read_excel(file_path, sheet_name='Critical Path Analysis')
        return main_df, cp_df
    except Exception as e:
        logging.error(f"Error loading data: {str(e)}")
        raise

def merge_data(main_df, cp_df):
    """Merge the main DataFrame with the critical path data."""
    merged_df = pd.merge(main_df, cp_df[['task_id', 'is_critical', 'total_float']], on='task_id', how='left')
    merged_df['is_critical'] = merged_df['is_critical'].fillna(False)
    return merged_df

def prepare_data_for_gantt(df, columns):
    """Prepare the merged DataFrame for Gantt chart creation."""
    gantt_columns = [columns['task_id'], columns['task_name'], columns['wbs_id'], columns['wbs_name'], 
                     columns['planned_duration'], columns['start_date'], columns['end_date'], 
                     'is_critical', 'total_float', columns['phys_complete_pct'], columns['task_type'], 
                     columns['task_code'], columns['rsrc_name'], columns['actv_code_name']]
    if 'predecessors' in columns:
        gantt_columns.append(columns['predecessors'])
    gantt_data = df[gantt_columns].copy()
    return gantt_data

def create_excel_gantt(gantt_data, columns, output_path):
    """Create an Excel file with a Gantt chart."""
    logging.info("Creating Excel Gantt chart")
    wb = Workbook()
    ws = wb.active
    ws.title = "Gantt Chart"

    # Write headers
    headers = ['Task ID', 'Task Name', 'WBS ID', 'WBS Name', 'Duration', 'Start Date', 'End Date', 'Is Critical', 'Total Float', 
               '% complete', 'type', 'task code', 'resource', 'alt WBS']
    if 'predecessors' in columns:
        headers.append('Predecessors')
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Write data
    for row, task in enumerate(gantt_data.itertuples(), start=2):
        ws.cell(row=row, column=1, value=getattr(task, columns['task_id']))
        ws.cell(row=row, column=2, value=getattr(task, columns['task_name']))
        ws.cell(row=row, column=3, value=getattr(task, columns['wbs_id']))
        ws.cell(row=row, column=4, value=getattr(task, columns['wbs_name']))
        ws.cell(row=row, column=5, value=getattr(task, columns['planned_duration']))
        ws.cell(row=row, column=6, value=getattr(task, columns['start_date']))
        ws.cell(row=row, column=7, value=getattr(task, columns['end_date']))
        ws.cell(row=row, column=8, value=getattr(task, 'is_critical'))
        ws.cell(row=row, column=9, value=getattr(task, 'total_float'))
        ws.cell(row=row, column=10, value=getattr(task, columns['phys_complete_pct']))
        ws.cell(row=row, column=11, value=getattr(task, columns['task_type']))
        ws.cell(row=row, column=12, value=getattr(task, columns['task_code']))
        ws.cell(row=row, column=13, value=getattr(task, columns['rsrc_name']))
        ws.cell(row=row, column=14, value=getattr(task, columns['actv_code_name']))
        if 'predecessors' in columns:
            ws.cell(row=row, column=15, value=getattr(task, columns['predecessors']))

    # Determine date range for Gantt chart
    start_date = min(gantt_data[columns['start_date']].min(), gantt_data[columns['end_date']].min())
    end_date = max(gantt_data[columns['start_date']].max(), gantt_data[columns['end_date']].max())
    date_range = (end_date - start_date).days + 1

    # Add date headers
    date_column_start = len(headers) + 1
    for i in range(date_range):
        date = start_date + timedelta(days=i)
        col = i + date_column_start
        ws.cell(row=1, column=col, value=date)
        ws.column_dimensions[get_column_letter(col)].width = 3

    # Add Gantt bars
    critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
    non_critical_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  # Blue

    for row, task in enumerate(gantt_data.itertuples(), start=2):
        task_start = getattr(task, columns['start_date'])
        task_end = getattr(task, columns['end_date'])
        is_critical = getattr(task, 'is_critical')
        
        if pd.isna(task_start) or pd.isna(task_end):
            continue
        
        start_col = (task_start - start_date).days + date_column_start
        end_col = (task_end - start_date).days + date_column_start
        fill = critical_fill if is_critical else non_critical_fill
        
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).fill = fill

    # Add legend
    ws.cell(row=len(gantt_data) + 3, column=1, value="Legend:")
    ws.cell(row=len(gantt_data) + 4, column=1, value="Critical Path")
    ws.cell(row=len(gantt_data) + 4, column=2).fill = critical_fill
    ws.cell(row=len(gantt_data) + 5, column=1, value="Non-Critical Path")
    ws.cell(row=len(gantt_data) + 5, column=2).fill = non_critical_fill

    # Apply some styling
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    wb.save(output_path)
    logging.info(f"Excel Gantt chart saved to: {output_path}")

def main(input_file_path, output_dir):
    """Main function to orchestrate the Gantt chart creation process."""
    logging.info(f"Starting script4 with input: {input_file_path}")

    try:
        main_df, cp_df = load_data(input_file_path)
        logging.info("Data loaded successfully")

        main_df = validate_and_clean_data(main_df)
        logging.info("Data validated and cleaned")

        columns = get_column_names(main_df)
        logging.info("Column names mapped")

        merged_df = merge_data(main_df, cp_df)
        logging.info("Data merged successfully")

        gantt_data = prepare_data_for_gantt(merged_df, columns)
        logging.info("Data prepared for Gantt chart")

        output_file_name = os.path.splitext(os.path.basename(input_file_path))[0] + "_gantt_chart.xlsx"
        output_file_path = os.path.join(output_dir, output_file_name)

        create_excel_gantt(gantt_data, columns, output_file_path)
        
        logging.info("Script4 completed successfully")
        return output_file_path
    
    except Exception as e:
        logging.error(f"An error occurred in script4: {str(e)}")
        raise

if __name__ == "__main__":
    print("This script is designed to be run as part of a larger pipeline.")
    print("Please use the orchestrator to run the full analysis.")